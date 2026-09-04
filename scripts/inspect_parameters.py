import os
import csv
import json
import pypdf
import openpyxl

PARAMS_DIR = r"c:\Users\Dell\Documents\sathi\parameters"

def inspect_csv(path):
    print(f"\n=================== CSV: {os.path.basename(path)} ===================")
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            print("Empty file")
            return
        
        print(f"Columns ({len(header)}): {header[:15]} ...")
        
        row_count = 0
        sample_rows = []
        for row in reader:
            row_count += 1
            if len(sample_rows) < 2:
                sample_rows.append(row)
        
        print(f"Total rows: {row_count}")
        for i, r in enumerate(sample_rows):
            # Print column-value pairs for first 10 cols
            preview = {header[j]: r[j] for j in range(min(len(header), len(r), 10))}
            print(f"Sample row {i+1} (first 10 cols): {preview}")

def inspect_pdf(path):
    print(f"\n=================== PDF: {os.path.basename(path)} ===================")
    try:
        reader = pypdf.PdfReader(path)
        print(f"Total pages: {len(reader.pages)}")
        first_page = reader.pages[0].extract_text()
        print("First page preview (first 400 chars):")
        print(first_page[:400].strip())
    except Exception as e:
        print(f"Error reading PDF: {e}")

def inspect_excel(path):
    print(f"\n=================== EXCEL: {os.path.basename(path)} ===================")
    try:
        # Load workbook read_only for large 82MB file
        wb = openpyxl.load_workbook(path, read_only=True)
        print("Sheet names:", wb.sheetnames)
        sheet = wb.active
        print(f"Active sheet: {sheet.title}")
        
        rows = sheet.iter_rows(values_only=True)
        header = None
        sample = []
        count = 0
        for r in rows:
            count += 1
            if count == 1:
                header = [str(c) if c is not None else '' for c in r]
                print(f"Header ({len(header)} cols): {header[:15]}")
            elif count <= 4:
                sample.append([str(c) if c is not None else '' for c in r])
            if count >= 100:
                break
        print(f"Sampled up to 100 rows. Row 2 preview: {dict(zip(header[:10], sample[0][:10])) if sample else 'None'}")
        wb.close()
    except Exception as e:
        print(f"Error reading Excel: {e}")

def main():
    files = sorted(os.listdir(PARAMS_DIR))
    for f in files:
        full_path = os.path.join(PARAMS_DIR, f)
        if f.endswith('.csv'):
            inspect_csv(full_path)
        elif f.endswith('.pdf'):
            inspect_pdf(full_path)
        elif f.endswith('.xlsx'):
            inspect_excel(full_path)

if __name__ == '__main__':
    main()
